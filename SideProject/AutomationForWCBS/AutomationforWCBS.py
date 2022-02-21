import threading
import openpyxl
from datetime import datetime
from openpyxl import workbook
import serial
import AutomationforWCBS_DB

count = 0 # Timer thread 동작을 위함
port = "COM3"
baud = 115200
ser = serial.Serial(port, baud)

excel_path = "" # 엑셀 저장 경로
excel_sheet_name = "" # 엑셀 저장 시트 이름
excution_count = 0

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb[excel_sheet_name]
ws = wb.active
ws.cell(row=5, column=3).value = datetime.today()
wb.save(excel_path)
wb.close()

def time_thread(): #Timer thread 동작을 위한 함수
    global count
    timer = threading.Timer(1, time_thread)
    timer.start()
    count += 1
    print(count)
    if count > 130: #80초 경과 시 종료
        ws.cell(row=6, column=3).value = datetime.today()  # End date 저장 위치
        ws.cell(row=8, column=3).value = excution_count # 총 진행 횟수 저장 위치
        wb.close()
        timer.cancel()
        ser.close()
        
time_thread()

if ser.isOpen() == False:
    ser.open()

while True:
    if ser.readable():
        res = ser.readline()
        try:
            res_str = res.decode("UTF-8")
            if "SW_NUMBER" in res_str:
                count = 0 # WCBS Init 되는 정보가 오면 Count 초기화
                excution_count += 1
                print(res_str)
                print("PASS, Count Init", "총 진행 횟수 : %d" %excution_count)

        except ValueError as e:
            print(e)
            continue