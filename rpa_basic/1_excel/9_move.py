from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
# 번호 (국어) 영어 수학  

# ws.move_range("B1:C11", rows=0, cols=1) # row는 그대로, cols만 1 만큼 이동 (이동하련느 데이터의 범위 설정, 이동할 거리 작성)
# ws["B1"].value = "국어" # B1 셀에 국어 입력

ws.move_range("C1:C11", rows=5, cols=-1) # 수학 기준 왼쪽으로 데이터 덮어 씀

wb.save("sample_korean.xlsx")