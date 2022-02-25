from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 생성
ws.title = "Mysheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ffffff"

ws1 = wb.create_sheet("Youtsheet") # 생성과 동시에 이름 변경
ws2 = wb.create_sheet("Newsheet", 1) # index로 sheet 위치 지정 가능

new_ws = wb["Newsheet"] # new_ws로 newsheet에 접근 가능
print(wb.sheetnames) # 모든 sheet 이름 확인
 
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws) # new_ws 복사
target.title = "Copied sheet"

wb.save("sample.xlsx")