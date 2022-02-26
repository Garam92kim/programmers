from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.merge_cells("B2:D2") # B2부터 D2까지 합침
ws["B2"].value = "Meged Cell"

wb.save("sample_merge.xlsx")
