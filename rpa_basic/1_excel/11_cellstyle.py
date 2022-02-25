from tkinter.ttk import Style
from turtle import left
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = load_workbook("sample.xlsx")
ws = wb.active

# 제목 값
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# 셀 스타일 적용
a1.font = Font(color="ff0000", italic=True, bold=True) # 글자 색 빨강색, 기울기, 두껍게 적용
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 폰트 Arianl 사용, 줄 긋기 사용
c1.font = Font(color="0000ff", size=20, underline="single") # 글자 크기 20, 밑줄 한줄 적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해 초록색 적용
for row in ws.rows:
    for cell in row:
        # 각 cell에 대해서 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center") # center, bottom, left, right 정렬 있음 / 모두 중앙정렬
        if cell.column == 1: # A 번호열은 제외
            continue
        
        if isinstance(cell.value, int) and cell.value > 90: # 이 인스턴스가 정수형고 90점보다 높을 때
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로 설정
            cell.font = Font(color="ff0000") # 폰트 색상도 변경
            
# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정

# # a열 너비 5로 설정
# ws.column_dimensions["A"].width = 5
# # 1행의 높이 50으로 설정
# ws.row_dimensions[1].height = 50

wb.save("sample_style.xlsx")