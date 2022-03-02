from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.append(("학번", "출석", "퀴즈1",	"퀴즈2", "중간고사", "기말고사", "프로젝트"))
scores = [(1, 10, 8, 5,	14,	26,	12),		
(2,	7,	3,	7,	15,	24,	18),	
(3,	9,	5,	8,	8,	12,	4),
(4,	7,	8,	7,	17,	21,	18),
(5,	7,	8,	7,	16,	25,	15),
(6,	3,	5,	8,	8,	17,	0),
(7,	4,	9,	10,	16,	27,	18),
(8,	6,	6,	6,	15,	19,	17),
(9,	10,	10,	9,	19,	30,	19),
(10, 9,	8, 8, 20, 25, 20)
]

for s in scores: # 기존 성적 데이터 넣기
    ws.append(s) # 한줄 씩 입력됨

for idx, cell in enumerate(ws["D"]): # D열만 가져와서 (퀴즈2) 값 업데이트, 제목열에는 값이 없으니 enumerate 사용
    if idx == 0:
        continue
    cell.value = 10
    
ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):  # 두번째 줄부터 값이 있으므로 idx 2부터 시작하도록 함 start=2 (1번 인덱스는 학번이므로 사용되지 않음)
    sum_val = sum(score[1:]) - score[3] + 10 # 총점 데이터로 사용될 것 (score 3) = 퀴즈2 에 해당하는 값은 제외하고 10을 더하도록 함
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx) # 총점에 해당하는 위치
    
    grade = None # 성적 정보
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    
    if score[1] < 5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade # i열에 성적 정보 추가    
wb.save("scores.xlsx")
