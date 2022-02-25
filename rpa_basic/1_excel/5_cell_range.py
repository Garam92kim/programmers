from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active



#1줄 씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) # 리스트 형태로도 넣을 수 있음
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"] # B 컬럼만 가져옴 (영어)
# for cell in col_B:
#     print(cell.value)
    
# col_range = ws["B:C"] # B부터 C까지 모든 컬럼 (영어, 수확) 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1] # 첫번째 row만 가져옴
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2번쨰 줄에서 6번째 줄까지 가져오기 slicing과 다르게 2와 6을 포함해서 가져옴 (slicing은 2부터 6 전까지)
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end = " ")
#     print()

# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end= " ") # 데이터만 가져옴
#         # print(cell.coordinate, end = " ") # 각 셀에대한 정보를 가져옴
#         xy = coordinate_from_string(cell.coordinate) # 좌표와 값을 분리해서 출력해줌
#         print(xy[0], end = " ")  # 좌표가 출력됨
#         print(xy[1], end= " ")
#         print(xy, end = " ")
#     print()

# 전체 Row를 가져오고 싶을때
# print(tuple(ws.rows)) # 한 줄씩 가져옴

# 전체 컬럼
# print(tuple(ws.columns)) # 한 열씩 가져옴

# for row in tuple(ws.rows):
#     print(row)
#     print(row[1].value) # 1번째 값  (모든 영어점수)

# for column in tuple(ws.columns):
#     print(column[0].value) # 제목만 출력

# for row in ws.iter_rows(): # 전체 row를 출력
#     print(row) # row[0], row[1] 이런식으로 사용 가능
    
# for column in ws.iter_cols():
#     print(column)

# iter 사용시 iter 안에 벙뮈를 지정할수 있음 (col과 row는 상하(cols), 좌우(rows) 로 데이터를 가져올 수 있음), 값을 지정하지 않으면 min부터 max 임
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3): # 1번 줄 부터 5번 줄 까지 진행, 2번째부터 3번째 열까지 출력
    # print(row[0].value, row[1].value) # 수학, 영어
    print(row)

wb.save("sample.xlsx")