from openpyxl import load_workbook # 파일 불러오기 역할

wb = load_workbook("sample.xlsx") # wb로 불러옴
ws = wb.active # 활성화된 sheet

# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row = x, column =y).value, end = " ")
#     print()

# cell 갯수를 모를 때 max_row, max_column 사용
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row = x, column =y).value, end = " ")
    print()