from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2): # 1줄 제목 스킵
    # 번호[0], 영어[1], 수학[2]
    if int(row[1].value) > 80:
        print(row[0].value, " 영어 고수!")

# 영어를 컴퓨터로 바꿔줌
for row in ws.iter_rows(max_row=1): # 1줄에대해서만 가져옴
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modi.xlsx")