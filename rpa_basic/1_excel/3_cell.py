from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Nadosheet"

ws["A1"] = 1 # A1셀에 1 입력
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 1 # A1셀에 1 입력
ws["B2"] = 2
ws["B3"] = 3

print(ws["A1"]) # A1 셀 정보 출력 (객체)
print(ws["A1"].value) # A1 셀 값 출력
print(ws["A10"].value) # 값 없으면 None 출력

print(ws.cell(row=1, column=1).value) # (=A1셀)

c = ws.cell(row=1, column=3, value=10) # 바로 10 입력
print(c.value)

from random import *
# 반복문 이용 랜덤 숫자 채우기
index = 1
for x in range(1, 11): # row
    for y in range(1, 11): # column
        # ws.cell(row=x, column=y, value=randint(0, 100))
        ws.cell(row = x, column = y, value = index) # 좌표 확인
        index += 1
        

wb.save("sample.xlsx")